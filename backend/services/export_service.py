"""
Servicio de exportación de facturas a Excel.
"""
from io import BytesIO
from typing import List
from datetime import datetime
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.models.database.models import Factura

logger = logging.getLogger(__name__)


class ExportService:
    """Servicio para exportar facturas a diferentes formatos"""
    
    def __init__(self):
        """Inicializar servicio"""
        logger.info("ExportService initialized")
    
    def export_facturas_to_excel(self, facturas: List[Factura], empresa_nombre: str = "") -> BytesIO:
        """
        Exportar lista de facturas a archivo Excel.
        
        Args:
            facturas: Lista de facturas a exportar
            empresa_nombre: Nombre de la empresa para el título
            
        Returns:
            BytesIO con el contenido del archivo Excel
        """
        logger.info(f"Exporting {len(facturas)} facturas to Excel")
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Facturas"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:N1')
        title_cell = ws['A1']
        title_cell.value = f"Reporte de Facturas - {empresa_nombre}" if empresa_nombre else "Reporte de Facturas"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Fecha de generación
        ws.merge_cells('A2:N2')
        date_cell = ws['A2']
        date_cell.value = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        date_cell.alignment = Alignment(horizontal='center')
        
        # Headers (fila 4)
        headers = [
            'ID', 'N° Factura', 'Fecha Emisión', 'Tipo',
            'Empresa Emisora', 'RUT Emisor', 'Domicilio Emisor',
            'Empresa Destinataria', 'RUT Destinatario', 'Domicilio Destinatario',
            'Monto Neto', 'IVA', 'Total', 'Estado'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        for row_num, factura in enumerate(facturas, start=5):
            # Tipo de factura
            tipo = ""
            if factura.tipo_factura:
                tipo = factura.tipo_factura.tipo
            
            # Fecha
            fecha_str = ""
            if factura.fecha_emision:
                if factura.fecha_emision.year != 1900:  # Fecha válida
                    fecha_str = factura.fecha_emision.strftime("%d/%m/%Y")
            
            row_data = [
                factura.id,
                factura.numero_factura if factura.numero_factura > 0 else "",
                fecha_str,
                tipo,
                factura.empresa_emisora or "",
                factura.rut_emisor or "",
                factura.domicilio_emisor or "",
                factura.empresa_destinataria or "",
                factura.rut_destinatario or "",
                factura.domicilio_destinatario or "",
                factura.monto_neto,
                factura.iva,
                factura.total,
                factura.status.value
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = thin_border
                
                # Formato de montos
                if col_num in [11, 12, 13]:  # Columnas de montos
                    if isinstance(value, int):
                        cell.number_format = '"$"#,##0'
                
                # Alineación
                if col_num in [1, 2]:  # ID y N° Factura
                    cell.alignment = Alignment(horizontal='center')
                elif col_num in [11, 12, 13]:  # Montos
                    cell.alignment = Alignment(horizontal='right')
        
        # Ajustar anchos de columna
        column_widths = {
            'A': 8,   # ID
            'B': 12,  # N° Factura
            'C': 15,  # Fecha
            'D': 12,  # Tipo
            'E': 30,  # Empresa Emisora
            'F': 15,  # RUT Emisor
            'G': 35,  # Domicilio Emisor
            'H': 30,  # Empresa Destinataria
            'I': 15,  # RUT Destinatario
            'J': 35,  # Domicilio Destinatario
            'K': 15,  # Monto Neto
            'L': 12,  # IVA
            'M': 15,  # Total
            'N': 12,  # Estado
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Fila de totales (si hay facturas)
        if facturas:
            total_row = len(facturas) + 5
            ws.cell(row=total_row, column=10).value = "TOTALES:"
            ws.cell(row=total_row, column=10).font = Font(bold=True)
            ws.cell(row=total_row, column=10).alignment = Alignment(horizontal='right')
            
            # Calcular totales
            total_neto = sum(f.monto_neto for f in facturas)
            total_iva = sum(f.iva for f in facturas)
            total_total = sum(f.total for f in facturas)
            
            ws.cell(row=total_row, column=11).value = total_neto
            ws.cell(row=total_row, column=11).number_format = '"$"#,##0'
            ws.cell(row=total_row, column=11).font = Font(bold=True)
            
            ws.cell(row=total_row, column=12).value = total_iva
            ws.cell(row=total_row, column=12).number_format = '"$"#,##0'
            ws.cell(row=total_row, column=12).font = Font(bold=True)
            
            ws.cell(row=total_row, column=13).value = total_total
            ws.cell(row=total_row, column=13).number_format = '"$"#,##0'
            ws.cell(row=total_row, column=13).font = Font(bold=True)
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Excel export completed. File size: {len(output.getvalue())} bytes")
        
        return output


# Singleton instance
export_service = ExportService()
